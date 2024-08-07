"""A vocabulary for declaratively describing how to import sheets.

Example:

    from tribune.sheet_import import Field, NestedItersSheetData, SheetImporter
    from tribune.sheet_import.parsers import (
        chain, parse_int, parse_text, parse_yes_no, validate_min
    )

    class MySheet(SheetImporter):
        fields = (
            Field('Name', 'name',    parse_text),
            Field('Age',  'age',     chain(parse_int, validate_min(1))),
            Field('Male', 'is_male', parse_yes_no),
        )

    sheet = NestedItersSheetData([
        ('Name',           'Age', 'Male'),
        ('Yuri Gagarin',   '34',  'Yes'),
        ('Alexei Leonov',  '81',  'Yes'),
        ('Amelia Earhart', '39',  'No'),
    ])
    records = MySheet().get_sheet_records(sheet)
"""

import abc
from collections import namedtuple
from io import BytesIO
from typing import BinaryIO
import warnings
from zipfile import BadZipFile

from blazeutils.spreadsheets import xlsx_to_reader, xlsx_to_strio
from xlsxwriter import Workbook


try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

import tribune

from ..utils import (
    column_letter,
    flatten,
    raises,
    split_every,
    unzip,
)


class SpreadsheetImportError(Exception):
    def __init__(self, errors):
        self.errors = errors

    def __str__(self):
        return f'<SpreadsheetImportError: {self.errors}>'


def xlrd_workbook_safe_open(file_handle):
    warnings.warn(
        'xlrd support is deprecated and will be removed in a future version',
        DeprecationWarning,
        stacklevel=2,
    )

    try:
        return xlrd.open_workbook(file_contents=file_handle.read())
    except xlrd.biffh.XLRDError as e:
        msg = str(e)
        # we don't want to screen all xlrd load errors, but a few common ones need to be trapped
        #   for user-friendliness
        if (
            'unsupported format' in msg.lower()
            or 'not supported' in msg.lower()
            or 'not a known type of workbook' in str(e).lower()
        ):
            raise SpreadsheetImportError(
                ['File format is not recognized. Please upload an Excel file (.xlsx or .xls).'],
            ) from e
        elif 'file size is 0 bytes' in msg.lower():
            raise SpreadsheetImportError(['File is empty.']) from e
        else:
            raise SpreadsheetImportError([msg]) from e


def workbook_safe_open(filename_or_filelike: str | BinaryIO):
    if not openpyxl:
        return xlrd_workbook_safe_open(filename_or_filelike)

    try:
        return openpyxl.load_workbook(filename_or_filelike, data_only=True)
    except (BadZipFile, InvalidFileException) as e:
        raise SpreadsheetImportError(
            ['File format is not recognized. Please upload an Excel file (.xlsx).'],
        ) from e


def normalize_text(x):
    """Normalizes a string by stripping whitespace and flattening case."""
    return str(x).strip().lower()


class Cell:
    """Represents a particular cell on a spreadsheet."""

    is_zero_based = False

    def __init__(self, row, column):
        """
        :param row: is a 0-based row index.
        :param column: is a 0-based column index.
        """
        self.row = row
        self.column = column

    def __str__(self):
        # If we're using openpyxl, column indices are one-based
        return column_letter(self.column, self.is_zero_based) + str(
            self.row + (1 if self.is_zero_based else 0),
        )


class XlrdCell(Cell):
    """Represents a single cell when reading sheets with xlrd, which uses
    0-based indices."""

    is_zero_based = True


class SheetDataBase(metaclass=abc.ABCMeta):
    """An ABC that describes the minimum necessary API for importing a sheet."""

    @abc.abstractmethod
    def __len__(self):
        raise NotImplementedError()

    @abc.abstractmethod
    def cell_value(self, cell):
        raise NotImplementedError()


class SheetData(SheetDataBase):
    def __init__(self, sheet):
        self.sheet = sheet

    def __len__(self):
        return self.sheet.max_row

    def cell_value(self, cell: Cell):
        try:
            return self.sheet.cell(cell.row, cell.column).value
        except IndexError:
            return ''

    @classmethod
    def from_nested_iters(cls, data):
        """Builds a SheetData from nested iterables.

        param data: is nested iterables representing data for each cell.
                    For example, `[('A1', 'B1'), ('A2', 'B2')]` represents two rows and two columns
                    of data.
        """
        workbook = Workbook(BytesIO())
        sheet = workbook.add_worksheet('test')

        for row, row_data in enumerate(data):
            for column, cell_data in enumerate(row_data):
                sheet.write(row, column, cell_data)

        workbook = openpyxl.load_workbook(xlsx_to_strio(workbook))
        sheet = workbook[workbook.sheetnames[0]]
        return SheetData(sheet)


class XlrdSheetData(SheetDataBase):
    def __init__(self, sheet):
        warnings.warn(
            'xlrd support is deprecated and will be removed in a future version',
            DeprecationWarning,
            stacklevel=2,
        )
        self.sheet = sheet

    def __len__(self):
        return self.sheet.nrows

    def cell_value(self, cell):
        try:
            return self.sheet.cell_value(cell.row, cell.column)
        except IndexError:
            return ''

    @classmethod
    def from_nested_iters(cls, data):
        """Builds an XlrdSheetData from nested iterables.

        param data: is nested iterables representing data for each cell.
                    For example, `[('A1', 'B1'), ('A2', 'B2')]` represents two rows and two columns
                    of data.
        """
        workbook = Workbook(BytesIO())
        sheet = workbook.add_worksheet('test')

        for row, row_data in enumerate(data):
            for column, cell_data in enumerate(row_data):
                sheet.write(row, column, cell_data)

        return XlrdSheetData(xlsx_to_reader(workbook).sheet_by_index(0))


class NestedItersSheetData(SheetDataBase):
    """Wraps nested iterables to provide the SheetData API."""

    def __init__(self, data):
        """
        param data: is nested iterables representing data for each cell.
                    For example, `[('A1', 'B1'), ('A2', 'B2')]` represents two rows and two columns
                    of data.
        """
        self.data = data

    def __len__(self):
        return len(self.data)

    def cell_value(self, cell):
        try:
            return self.data[cell.row][cell.column]
        except IndexError:
            return ''


ErrorOrValue = namedtuple('ErrorOrValue', ['error', 'value'])


def aggregate_error_or_value(error_or_values):
    """Aggregates a list of ErrorOrValue tuples into a single ErrorOrValue (very much like `unzip`).
    If there any errors the resulting ErrorOrValue will only have list of errors.
    If there are no errors the resulting ErrorOrValue will have a list of values.
    """
    raw_errors, values = list(unzip(error_or_values)) or ([], [])
    errors = [error for error in flatten(raw_errors) if error is not None]
    return ErrorOrValue(errors, None) if errors else ErrorOrValue([], values)


Field = namedtuple('Field', ['label', 'field', 'parser'])


class SheetImporter:
    """A base class for defining how to parse a single spreadsheet."""

    data_start_row = 2  # openpyxl indices are 1-based
    data_start_col = 1
    is_zero_based = False
    cell_cls = Cell
    fields = ()  # Ordered iterable of Field tuples

    def __init__(self, fields=None, data_start_row=None):
        """
        :param fields: is an ordered iterable of Field tuples. If not provided, the class-level
                       value will be used.
        :param data_start_row: is the 0-based index where the first row of data begins. If not
                               provided, the class-level value will be used.
        """
        self.fields = fields or self.fields
        self.data_start_row = data_start_row or self.data_start_row

    def _get_header_errors(self, sheet):
        """Builds a list of errors regarding the sheet header. The list will be empty if there are
        no errors."""
        header = self.data_start_row - 1
        return [
            f'Expected "{field.label}" in header cell {self.cell_cls(header, col)}.'
            for col, field in enumerate(self.fields, self.data_start_col)
            if normalize_text(sheet.cell_value(self.cell_cls(header, col)))
            != normalize_text(field.label)
        ]

    @staticmethod
    def _parse_cell(field, cell, value):
        """Parses a cell using the given field definition. The result is an ErrorOrValue tuple.

        :param field: is a Field tuple which defines the field.
        :param cell: is a Cell which contains both row and column indexes.
        :param value: is the raw value at that cell in the sheet.
        """
        try:
            return ErrorOrValue([], field.parser(value))
        except SpreadsheetImportError as e:
            contextualized_errors = [f'Unexpected value in cell {cell}: {msg}' for msg in e.errors]
            return ErrorOrValue(contextualized_errors, None)

    def _parse_row(self, sheet, row):
        """Parses all the columns in a sheet at a given row index and returns aggregated errors or
        values (for every column) as an ErrorOrValue tuple."""
        col_data = [
            self._parse_cell(
                field,
                self.cell_cls(row, col),
                sheet.cell_value(self.cell_cls(row, col)),
            )
            for col, field in enumerate(self.fields, self.data_start_col)
        ]
        errors, values = aggregate_error_or_value(col_data)
        return (
            ErrorOrValue(errors, None) if errors else ErrorOrValue([], self.row_to_record(values))
        )

    def _parse_sheet_data(self, sheet):
        """Parses only the data rows from a sheet and returns aggregated errors or valuse as an
        ErrorOrValue tuple."""
        return [
            self._parse_row(sheet, row)
            for row in range(
                self.data_start_row,
                len(sheet) + (1 if not self.is_zero_based else 0),
            )
        ]

    def _parse_sheet(self, sheet):
        """Parses a sheet (including the headers) and returns aggregated errors or values as an
        ErrorOrValue tuple."""
        header_errors = self._get_header_errors(sheet)
        if header_errors:
            return [ErrorOrValue(header_errors, None)]
        return self._parse_sheet_data(sheet)

    def row_to_record(self, values):
        """Converts a list of values from a row into a dictionary where each value is keyed
        on its column name from `self.fields`."""
        return self.modify_record(
            {field.field: v for field, v in zip(self.fields, values, strict=True)},
        )

    def modify_record(self, record):
        """Override to add/remove/change fields in the record (dictionary of column to value).

        :returns: the record with any modifications applied or None to filter out the record
                  entirely.
        """
        return record

    def get_sheet_records(self, sheet):
        """Parses a sheet and returns an iterator of rows where each row is a dictionary of data
        values keyed by their column name. If errors occurred during the parse, this aggregates them
        and raises them in `SpreadsheetImportError`."""
        errors_or_values = self._parse_sheet(sheet)
        errors, records = aggregate_error_or_value(errors_or_values)
        # Filter out empty records.
        return filter(None, records) if not errors else raises(SpreadsheetImportError(errors))

    def get_sheet_records_batched(self, sheet, batch_size):
        """Like `get_sheet_records` but returns a generator that splits the result into batches."""
        return split_every(batch_size, self.get_sheet_records(sheet))

    def as_report_sheet(self):
        """Returns a ReportSheet class which mirrors this SheetImporter."""

        class NewReportSheet(tribune.ReportSheet):
            pre_data_rows = self.data_start_row

            for ___field in self.fields:  # This iteration variable gets left behind on the class!
                tribune.LabeledColumn(___field.label)

            def fetch_records(self):
                return []

        return NewReportSheet


class XlrdSheetImporter(SheetImporter):
    """A base class for defining how to parse a single spreadsheet."""

    data_start_row = 1  # xlrd rows are 0-based
    data_start_col = 0
    cell_cls = XlrdCell
    is_zero_based = True
